"""Excel（.xlsx / .xls）用例导入解析器。

使用 openpyxl 处理 xlsx，不依赖 pandas 以保持轻量。
若 openpyxl 不可用，回退到 csv 方式（用户需转为 CSV 后重新上传）。
"""

from __future__ import annotations

import io
import logging

logger = logging.getLogger(__name__)


def parse_excel(data: bytes, *, sheet_name: str | None = None) -> list[dict]:
    """解析 Excel 字节流，返回统一 list[dict] 格式。

    Args:
        data: Excel 文件字节内容
        sheet_name: 指定工作表名（None 时使用第一个）

    Returns:
        list[dict]: 每行数据，key 为表头列名
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.error("openpyxl 未安装，无法解析 Excel 文件")
        raise RuntimeError("Excel 解析需要 openpyxl，请安装: uv add openpyxl") from None

    wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    if ws is None:
        return []

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []

    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(header_row)]

    rows: list[dict] = []
    for i, row in enumerate(rows_iter, 1):
        if all(cell is None for cell in row):
            continue
        record: dict = {"_row_number": i}
        for col_name, cell_val in zip(headers, row, strict=False):
            record[col_name] = str(cell_val).strip() if cell_val is not None else ""
        rows.append(record)

    wb.close()
    logger.info("Excel 解析完成，共 %d 行", len(rows))
    return rows
